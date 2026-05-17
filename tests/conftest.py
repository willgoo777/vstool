"""程序化生成测试用 fixtures（成对 xlsx + docx zip）。

之所以不在仓库里放二进制：体积大、diff 噪声大、不便审查。
"""
from __future__ import annotations

import os
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# 确保 src 在路径上（不依赖 pip install -e .）
ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
sys.path.insert(0, str(SRC))


# ---------- Excel fixtures ----------

def _make_wb_values(values: dict[str, list[list]]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in values.items():
        ws = wb.create_sheet(sheet_name)
        for r, row in enumerate(rows, 1):
            for c, v in enumerate(row, 1):
                ws.cell(row=r, column=c, value=v)
    return wb


# ---------- Docx fixture（最小可读 .docx） ----------

_DOCX_DOC_TEMPLATE = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>
{paragraphs}
<w:sectPr/></w:body></w:document>"""

_DOCX_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""

_DOCX_ROOT_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""


def _make_docx(path: Path, paragraphs: list[str]) -> None:
    body = "\n".join(
        f"<w:p><w:r><w:t xml:space='preserve'>{p}</w:t></w:r></w:p>"
        for p in paragraphs
    )
    doc_xml = _DOCX_DOC_TEMPLATE.format(paragraphs=body)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", _DOCX_CONTENT_TYPES)
        z.writestr("_rels/.rels", _DOCX_ROOT_RELS)
        z.writestr("word/document.xml", doc_xml)


# ---------- Fixture 工厂 ----------

@pytest.fixture
def workdir(tmp_path: Path) -> Path:
    return tmp_path


@pytest.fixture
def paired_dirs(workdir: Path) -> tuple[Path, Path, Path]:
    """构造 A、B、output 三个目录并放入若干成对/孤立文件。

    布局：
      A/identical.xlsx           # 与 B 完全相同
      B/identical.xlsx
      A/values.xlsx              # 单元格值差异
      B/values.xlsx
      A/sub/nested.xlsx          # 子文件夹差异
      B/sub/nested.xlsx
      A/only_a.xlsx              # 仅 A 有
      B/only_b.xlsx              # 仅 B 有
      A/letter.docx + B/letter.docx
    """
    a = workdir / "A"
    b = workdir / "B"
    out = workdir / "OUT"
    for d in (a, b, out):
        d.mkdir()

    # identical
    wb1 = _make_wb_values({"Sheet1": [["x", 1], ["y", 2]]})
    wb1.save(a / "identical.xlsx")
    wb1.save(b / "identical.xlsx")

    # values
    wb2a = _make_wb_values({"Sheet1": [["x", 1], ["y", 2]]})
    wb2b = _make_wb_values({"Sheet1": [["x", 1], ["y", 99]]})  # B[B2] 不同
    wb2a.save(a / "values.xlsx")
    wb2b.save(b / "values.xlsx")

    # 子文件夹
    (a / "sub").mkdir()
    (b / "sub").mkdir()
    wb3a = _make_wb_values({"Sheet1": [[1, 2], [3, 4]]})
    wb3b = _make_wb_values({"Sheet1": [[1, 2], [3, 5]]})  # B[B2] 4→5
    wb3a.save(a / "sub" / "nested.xlsx")
    wb3b.save(b / "sub" / "nested.xlsx")

    # 仅 A / 仅 B
    _make_wb_values({"Sheet1": [["x"]]}).save(a / "only_a.xlsx")
    _make_wb_values({"Sheet1": [["x"]]}).save(b / "only_b.xlsx")

    # docx 成对
    _make_docx(a / "letter.docx", ["你好，世界。", "这是第二段。"])
    _make_docx(b / "letter.docx", ["你好，世界。", "这是改过的第二段。"])

    return a, b, out


@pytest.fixture
def format_diff_xlsx(workdir: Path) -> tuple[Path, Path]:
    """仅格式不同（字体/填充）的成对 xlsx。"""
    a = workdir / "a_fmt.xlsx"
    b = workdir / "b_fmt.xlsx"
    wba = _make_wb_values({"Sheet1": [["hello"]]})
    wbb = _make_wb_values({"Sheet1": [["hello"]]})
    wbb["Sheet1"].cell(1, 1).font = Font(bold=True)
    wbb["Sheet1"].cell(1, 1).fill = PatternFill("solid", fgColor="FFFFFF00")
    wba.save(a)
    wbb.save(b)
    return a, b


@pytest.fixture
def formula_only_diff_xlsx(workdir: Path) -> tuple[Path, Path]:
    """公式不同但缓存值相同。"""
    a = workdir / "a_form.xlsx"
    b = workdir / "b_form.xlsx"
    wba = Workbook(); wba.remove(wba.active); ws = wba.create_sheet("S")
    ws["A1"] = 3
    ws["A2"] = "=1+2"          # 值 3
    wbb = Workbook(); wbb.remove(wbb.active); ws2 = wbb.create_sheet("S")
    ws2["A1"] = 3
    ws2["A2"] = "=2+1"         # 公式不同，值仍为 3
    wba.save(a); wbb.save(b)
    return a, b


@pytest.fixture
def structure_diff_xlsx(workdir: Path) -> tuple[Path, Path]:
    """B 多一个 sheet。"""
    a = workdir / "a_struct.xlsx"
    b = workdir / "b_struct.xlsx"
    _make_wb_values({"S1": [[1]]}).save(a)
    _make_wb_values({"S1": [[1]], "S2": [[2]]}).save(b)
    return a, b
