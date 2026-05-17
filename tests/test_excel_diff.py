from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from vstool.excel_diff import (
    FORMAT_DIFF,
    FORMULA_DIFF,
    SHEET_ADDED,
    VALUE_DIFF,
    diff_workbooks,
)
from vstool.i18n import T


def test_identical_files_have_no_changes(workdir: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active); ws = wb.create_sheet("S")
    ws["A1"] = 1; ws["B1"] = 2
    a = workdir / "a.xlsx"; b = workdir / "b.xlsx"; out = workdir / "out.xlsx"
    wb.save(a); wb.save(b)

    result = diff_workbooks(a, b, out)
    assert result.has_changes is False
    assert out.exists()
    # 输出至少有总览和明细两个 sheet
    wb_out = load_workbook(out)
    assert T["excel_sheet_summary"] in wb_out.sheetnames
    assert T["excel_sheet_details"] in wb_out.sheetnames


def test_value_diff_detected_and_colored(paired_dirs) -> None:
    a_dir, b_dir, out = paired_dirs
    out_file = out / "values_diff.xlsx"
    result = diff_workbooks(a_dir / "values.xlsx", b_dir / "values.xlsx", out_file)

    value_diffs = [d for d in result.cell_diffs if d.change_type == VALUE_DIFF]
    assert len(value_diffs) == 1
    d = value_diffs[0]
    assert d.sheet == "Sheet1"
    assert d.cell == "B2"
    assert d.a_value == 2 and d.b_value == 99

    # 可视化 sheet 存在且 B2 被染色
    wb_out = load_workbook(out_file)
    assert "S_Sheet1" in wb_out.sheetnames
    cell = wb_out["S_Sheet1"]["B2"]
    assert cell.fill.patternType == "solid"


def test_format_only_diff(format_diff_xlsx, workdir: Path) -> None:
    a, b = format_diff_xlsx
    out_file = workdir / "fmt_diff.xlsx"
    result = diff_workbooks(a, b, out_file)
    types = {d.change_type for d in result.cell_diffs}
    assert FORMAT_DIFF in types
    assert VALUE_DIFF not in types


def test_formula_only_diff(formula_only_diff_xlsx, workdir: Path) -> None:
    a, b = formula_only_diff_xlsx
    out_file = workdir / "form_diff.xlsx"
    result = diff_workbooks(a, b, out_file)
    formula_diffs = [d for d in result.cell_diffs if d.change_type == FORMULA_DIFF]
    # 注意：openpyxl 的 data_only=True 在文件未经 Excel 打开保存时缓存值可能为 None；
    # 此处我们手动写了 cell 值为 3 + 公式串，data_only 读到 None → 走 cached_missing
    # 路径，最终判为 formula diff（不再当值差）。
    assert any(d.cell == "A2" for d in formula_diffs)


def test_structure_diff_sheet_added(structure_diff_xlsx, workdir: Path) -> None:
    a, b = structure_diff_xlsx
    out_file = workdir / "struct_diff.xlsx"
    result = diff_workbooks(a, b, out_file)
    sheet_adds = [s for s in result.struct_diffs if s.change_type == SHEET_ADDED]
    assert [s.sheet for s in sheet_adds] == ["S2"]
