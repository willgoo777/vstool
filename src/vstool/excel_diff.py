"""Excel 四维度对比：值 / 公式 / 格式 / 结构。

输出单个 .xlsx：
- 00_总览：每 sheet 改动计数、增删 sheet、合并区域差异、命名区间差异
- S_<sheet>：复制 B 的工作表，按差异类型上色（红/黄/蓝），新增行/列表头标绿
- 99_差异明细：表格式明细，便于筛选

设计取舍：
- 用 openpyxl 同时打开 data_only=True（计算值）与 data_only=False（公式）两份副本
- 合并单元格只比较锚点（左上角）；其他位置跳过避免噪音
- theme/indexed 颜色直接比对 .rgb 不可靠，回退到结构化元组
- data_only 读到 None 但公式存在 → 缓存值缺失，不当差异报
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Color, Fill, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import (
    COLOR_ADDED_HEADER,
    COLOR_FORMAT_DIFF,
    COLOR_FORMULA_DIFF,
    COLOR_SUMMARY_HEADER,
    COLOR_VALUE_DIFF,
)
from .i18n import T


# ---------- 数据结构 ----------

VALUE_DIFF = "value"
FORMULA_DIFF = "formula"
FORMAT_DIFF = "format"
A_ONLY = "a_only"
B_ONLY = "b_only"
SHEET_ADDED = "sheet_added"
SHEET_REMOVED = "sheet_removed"
MERGED_DIFF = "merged"
NAMED_RANGE_DIFF = "named_range"

_CT_LABEL = {
    VALUE_DIFF: T["excel_ct_value"],
    FORMULA_DIFF: T["excel_ct_formula"],
    FORMAT_DIFF: T["excel_ct_format"],
    A_ONLY: T["excel_ct_a_only"],
    B_ONLY: T["excel_ct_b_only"],
    SHEET_ADDED: T["excel_ct_sheet_added"],
    SHEET_REMOVED: T["excel_ct_sheet_removed"],
    MERGED_DIFF: T["excel_ct_merged_diff"],
    NAMED_RANGE_DIFF: T["excel_ct_named_range"],
}


@dataclass
class CellDiff:
    sheet: str
    cell: str          # A1 形式
    change_type: str
    a_value: Any = None
    b_value: Any = None
    a_formula: str | None = None
    b_formula: str | None = None
    a_format_sig: str = ""
    b_format_sig: str = ""
    note: str = ""


@dataclass
class SheetStructDiff:
    change_type: str   # SHEET_ADDED / SHEET_REMOVED / MERGED_DIFF / NAMED_RANGE_DIFF
    sheet: str = ""
    detail: str = ""


@dataclass
class ExcelDiffResult:
    output_path: Path
    cell_diffs: list[CellDiff] = field(default_factory=list)
    struct_diffs: list[SheetStructDiff] = field(default_factory=list)
    has_changes: bool = False


# ---------- 格式签名 ----------

def _color_key(color: Color | None) -> tuple:
    if color is None:
        return ()
    return (
        getattr(color, "type", None),
        getattr(color, "rgb", None),
        getattr(color, "theme", None),
        getattr(color, "indexed", None),
        getattr(color, "tint", None),
    )


def _font_key(font: Font | None) -> tuple:
    if font is None:
        return ()
    return (
        font.name, font.size, bool(font.bold), bool(font.italic),
        bool(font.underline), bool(font.strike),
        _color_key(font.color),
    )


def _fill_key(fill: Fill | None) -> tuple:
    if fill is None:
        return ()
    if isinstance(fill, PatternFill):
        return ("pattern", fill.patternType, _color_key(fill.fgColor), _color_key(fill.bgColor))
    return (type(fill).__name__,)


def _align_key(a: Alignment | None) -> tuple:
    if a is None:
        return ()
    return (a.horizontal, a.vertical, bool(a.wrap_text), a.text_rotation, a.indent)


def _border_key(b: Border | None) -> tuple:
    if b is None:
        return ()
    sides = (b.left, b.right, b.top, b.bottom)
    return tuple(
        (s.style if s else None, _color_key(s.color) if s else ())
        for s in sides
    )


def _format_signature(cell: Cell) -> str:
    key = (
        _font_key(cell.font),
        _fill_key(cell.fill),
        _align_key(cell.alignment),
        _border_key(cell.border),
        cell.number_format,
    )
    raw = repr(key).encode("utf-8")
    return hashlib.md5(raw).hexdigest()[:12]


# ---------- 比对核心 ----------

def _safe_load(path: Path, data_only: bool) -> Workbook:
    return load_workbook(path, data_only=data_only, read_only=False, keep_vba=False)


def _is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _merged_anchors(ws: Worksheet) -> set[str]:
    """返回所有合并区域的左上角单元格坐标集合（用于跳过非锚点）。"""
    anchors: set[str] = set()
    non_anchor: set[str] = set()
    for rng in ws.merged_cells.ranges:
        rows = list(ws[rng.coord])
        if not rows:
            continue
        first = rows[0][0].coordinate
        anchors.add(first)
        for row in rows:
            for c in row:
                if c.coordinate != first:
                    non_anchor.add(c.coordinate)
    return non_anchor  # 返回需要跳过的非锚点集合


def _diff_one_sheet(
    name: str,
    a_val: Worksheet, a_form: Worksheet,
    b_val: Worksheet, b_form: Worksheet,
) -> tuple[list[CellDiff], list[SheetStructDiff]]:
    cell_diffs: list[CellDiff] = []
    struct_diffs: list[SheetStructDiff] = []

    # 合并区域差异
    a_merged = {str(r) for r in a_val.merged_cells.ranges}
    b_merged = {str(r) for r in b_val.merged_cells.ranges}
    if a_merged != b_merged:
        only_a = sorted(a_merged - b_merged)
        only_b = sorted(b_merged - a_merged)
        detail = ""
        if only_a:
            detail += f"仅 A 有：{', '.join(only_a)}；"
        if only_b:
            detail += f"仅 B 有：{', '.join(only_b)}"
        struct_diffs.append(SheetStructDiff(change_type=MERGED_DIFF, sheet=name, detail=detail))

    # 单元格遍历范围：合并 A、B 的最大行列
    max_row = max(a_val.max_row or 0, b_val.max_row or 0)
    max_col = max(a_val.max_column or 0, b_val.max_column or 0)
    if max_row == 0 or max_col == 0:
        return cell_diffs, struct_diffs

    skip_a = _merged_anchors(a_val)
    skip_b = _merged_anchors(b_val)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            coord = f"{get_column_letter(col)}{row}"
            if coord in skip_a or coord in skip_b:
                continue

            ac_val = a_val.cell(row=row, column=col)
            bc_val = b_val.cell(row=row, column=col)
            ac_form = a_form.cell(row=row, column=col)
            bc_form = b_form.cell(row=row, column=col)

            a_v = ac_val.value
            b_v = bc_val.value
            a_f = ac_form.value if _is_formula(ac_form.value) else None
            b_f = bc_form.value if _is_formula(bc_form.value) else None

            a_sig = _format_signature(ac_val)
            b_sig = _format_signature(bc_val)

            a_empty = a_v is None and a_f is None
            b_empty = b_v is None and b_f is None

            # 仅 A 有
            if not a_empty and b_empty:
                cell_diffs.append(CellDiff(
                    sheet=name, cell=coord, change_type=A_ONLY,
                    a_value=a_v, b_value=None,
                    a_formula=a_f, b_formula=None,
                    a_format_sig=a_sig, b_format_sig="",
                    note=T["excel_note_a_only"],
                ))
                continue
            # 仅 B 有
            if a_empty and not b_empty:
                cell_diffs.append(CellDiff(
                    sheet=name, cell=coord, change_type=B_ONLY,
                    a_value=None, b_value=b_v,
                    a_formula=None, b_formula=b_f,
                    a_format_sig="", b_format_sig=b_sig,
                    note=T["excel_note_b_only"],
                ))
                continue
            if a_empty and b_empty:
                continue

            # 公式差异（即使计算值相同也单独标）
            formula_changed = (a_f or "") != (b_f or "")
            # 值差异：若任一侧有公式但 data_only 读到 None，标注缓存缺失而不报值差
            cached_missing = False
            value_changed = False
            if a_v != b_v:
                if (a_f and a_v is None) or (b_f and b_v is None):
                    cached_missing = True
                else:
                    value_changed = True

            format_changed = a_sig != b_sig

            if value_changed:
                cell_diffs.append(CellDiff(
                    sheet=name, cell=coord, change_type=VALUE_DIFF,
                    a_value=a_v, b_value=b_v,
                    a_formula=a_f, b_formula=b_f,
                    a_format_sig=a_sig, b_format_sig=b_sig,
                ))
            elif formula_changed:
                cell_diffs.append(CellDiff(
                    sheet=name, cell=coord, change_type=FORMULA_DIFF,
                    a_value=a_v, b_value=b_v,
                    a_formula=a_f, b_formula=b_f,
                    a_format_sig=a_sig, b_format_sig=b_sig,
                    note=T["excel_note_cached_none"] if cached_missing else "",
                ))
            elif format_changed:
                cell_diffs.append(CellDiff(
                    sheet=name, cell=coord, change_type=FORMAT_DIFF,
                    a_value=a_v, b_value=b_v,
                    a_formula=a_f, b_formula=b_f,
                    a_format_sig=a_sig, b_format_sig=b_sig,
                ))

    return cell_diffs, struct_diffs


def _named_range_diff(
    wb_a: Workbook, wb_b: Workbook,
) -> list[SheetStructDiff]:
    a_names = {name: defn.value for name, defn in wb_a.defined_names.items()}
    b_names = {name: defn.value for name, defn in wb_b.defined_names.items()}
    diffs: list[SheetStructDiff] = []
    for k in sorted(set(a_names) | set(b_names)):
        if k not in a_names:
            diffs.append(SheetStructDiff(NAMED_RANGE_DIFF, "", f"B 新增：{k} = {b_names[k]}"))
        elif k not in b_names:
            diffs.append(SheetStructDiff(NAMED_RANGE_DIFF, "", f"B 删除：{k}（A 中 = {a_names[k]}）"))
        elif a_names[k] != b_names[k]:
            diffs.append(SheetStructDiff(
                NAMED_RANGE_DIFF, "",
                f"{k}：A = {a_names[k]} → B = {b_names[k]}"))
    return diffs


# ---------- 输出写入 ----------

def _color_for(ct: str) -> str | None:
    return {
        VALUE_DIFF: COLOR_VALUE_DIFF,
        FORMULA_DIFF: COLOR_FORMULA_DIFF,
        FORMAT_DIFF: COLOR_FORMAT_DIFF,
        A_ONLY: COLOR_VALUE_DIFF,
        B_ONLY: COLOR_ADDED_HEADER,
    }.get(ct)


def _write_summary_sheet(
    wb_out: Workbook,
    cell_diffs: list[CellDiff],
    struct_diffs: list[SheetStructDiff],
) -> None:
    ws = wb_out.create_sheet(T["excel_sheet_summary"], 0)
    headers = [
        T["excel_summary_h_sheet"],
        T["excel_summary_h_change_type"],
        T["excel_summary_h_count"],
        T["excel_summary_h_note"],
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_HEADER)

    # 按 (sheet, change_type) 聚合 cell_diffs
    counts: dict[tuple[str, str], int] = {}
    for d in cell_diffs:
        counts[(d.sheet, d.change_type)] = counts.get((d.sheet, d.change_type), 0) + 1
    for (sheet, ct), n in sorted(counts.items()):
        ws.append([sheet, _CT_LABEL.get(ct, ct), n, ""])

    # 结构差异
    for sd in struct_diffs:
        ws.append([sd.sheet, _CT_LABEL.get(sd.change_type, sd.change_type), 1, sd.detail])

    # 列宽
    widths = [20, 14, 8, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_details_sheet(wb_out: Workbook, cell_diffs: list[CellDiff]) -> None:
    ws = wb_out.create_sheet(T["excel_sheet_details"])
    headers = T["excel_details_headers"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_HEADER)
    for d in cell_diffs:
        ws.append([
            d.sheet, d.cell, _CT_LABEL.get(d.change_type, d.change_type),
            _stringify(d.a_value), _stringify(d.b_value),
            d.a_formula or "", d.b_formula or "",
            d.a_format_sig, d.b_format_sig,
        ])
    widths = [18, 8, 12, 22, 22, 24, 24, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return str(v)
    return str(v)


def _write_visual_sheet(
    wb_out: Workbook,
    sheet_name: str,
    b_ws: Worksheet,
    diffs_in_sheet: list[CellDiff],
) -> None:
    """复制 B 的 sheet 内容（值，不带原格式以保持简洁），按差异类型上色。"""
    out = wb_out.create_sheet(f"S_{sheet_name}"[:31])
    max_row = b_ws.max_row or 0
    max_col = b_ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = b_ws.cell(row=r, column=c).value
            if v is None:
                continue
            out.cell(row=r, column=c, value=v)

    # 上色 + 批注
    diff_index = {d.cell: d for d in diffs_in_sheet}
    for coord, d in diff_index.items():
        color = _color_for(d.change_type)
        if not color:
            continue
        cell = out[coord]
        cell.fill = PatternFill("solid", fgColor=color)
        comment_lines = [f"类型：{_CT_LABEL.get(d.change_type, d.change_type)}"]
        if d.a_value is not None or d.a_formula:
            comment_lines.append(f"A 值：{_stringify(d.a_value)}")
            if d.a_formula:
                comment_lines.append(f"A 公式：{d.a_formula}")
        if d.b_formula and d.b_formula != d.a_formula:
            comment_lines.append(f"B 公式：{d.b_formula}")
        if d.note:
            comment_lines.append(f"备注：{d.note}")
        cell.comment = Comment("\n".join(comment_lines), "vstool")

    # 大致列宽
    for col in range(1, max_col + 1):
        out.column_dimensions[get_column_letter(col)].width = 16


# ---------- 对外入口 ----------

def diff_workbooks(a_path: Path, b_path: Path, out_path: Path) -> ExcelDiffResult:
    """对比 A 与 B 两个 xlsx，把综合差异写入 out_path。"""
    wb_a_val = _safe_load(a_path, data_only=True)
    wb_a_form = _safe_load(a_path, data_only=False)
    wb_b_val = _safe_load(b_path, data_only=True)
    wb_b_form = _safe_load(b_path, data_only=False)

    cell_diffs: list[CellDiff] = []
    struct_diffs: list[SheetStructDiff] = []

    a_sheets = wb_a_val.sheetnames
    b_sheets = wb_b_val.sheetnames
    common = [s for s in b_sheets if s in a_sheets]

    for s in b_sheets:
        if s not in a_sheets:
            struct_diffs.append(SheetStructDiff(SHEET_ADDED, s, ""))
    for s in a_sheets:
        if s not in b_sheets:
            struct_diffs.append(SheetStructDiff(SHEET_REMOVED, s, ""))

    for s in common:
        cds, sds = _diff_one_sheet(s,
                                   wb_a_val[s], wb_a_form[s],
                                   wb_b_val[s], wb_b_form[s])
        cell_diffs.extend(cds)
        struct_diffs.extend(sds)

    struct_diffs.extend(_named_range_diff(wb_a_form, wb_b_form))

    # 写出
    wb_out = Workbook()
    # 删默认 sheet
    default = wb_out.active
    wb_out.remove(default)

    _write_summary_sheet(wb_out, cell_diffs, struct_diffs)

    # 仅为有差异的 sheet 生成可视化
    sheets_with_diffs = sorted({d.sheet for d in cell_diffs} & set(b_sheets))
    for s in sheets_with_diffs:
        per_sheet = [d for d in cell_diffs if d.sheet == s]
        _write_visual_sheet(wb_out, s, wb_b_val[s], per_sheet)

    _write_details_sheet(wb_out, cell_diffs)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)

    return ExcelDiffResult(
        output_path=out_path,
        cell_diffs=cell_diffs,
        struct_diffs=struct_diffs,
        has_changes=bool(cell_diffs or struct_diffs),
    )
